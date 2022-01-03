import openpyxl


class ExcelUtils:
    def __init__(self, file_path, sheetName='Sheet1'):
        self.file_path = file_path
        self.sheetName = sheetName
        self.workbook = openpyxl.load_workbook(file_path)
        # self.sheet = self.workbook[sheetName]
        self.sheet = self.workbook.active
        self.row_count = self.sheet.max_row
        self.column_count = self.sheet.max_column

    def read_cell(self, rowNum, columnNum):
        return self.sheet.cell(rowNum, columnNum).value

    def write_to_cell(self, rowNum, columnNum, value):
        self.sheet.cell(rowNum, columnNum).value = value
        self.workbook.save(self.file_path)
        # self.sheet.

# s = ExcelUtils('.//TestData//TestData.xlsx', 'Sheet1')
